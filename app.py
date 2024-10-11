from flask import Flask, request, render_template, send_file
import re
import openpyxl
import csv
import io
from openpyxl import Workbook

app = Flask(__name__)

# Hàm để trích xuất các liên kết từ văn bản
def extract_links(text):
    pattern = r'(https?://[^\s]+)'
    links = re.findall(pattern, text)
    return links

# Hàm để đọc file CSV chứa liên kết gốc và liên kết thay thế
def read_csv(file):
    replacements = {}
    reader = csv.reader(io.StringIO(file.read().decode('utf-8')))
    for row in reader:
        if len(row) > 6:  # Đảm bảo có đủ cột A và G
            original_link = row[0]  # Liên kết gốc ở cột A
            replacement_link = row[6]  # Liên kết thay thế ở cột G
            replacements[original_link] = replacement_link
    return replacements

# Hàm để tạo file Excel trong bộ nhớ
def create_excel(links):
    wb = Workbook()
    ws = wb.active
    ws.title = "Links"
    
    # Ghi các nhãn cho hàng số 1
    ws['A1'] = 'Liên kết gốc'
    ws['B1'] = 'Sub_id1'
    ws['C1'] = 'Sub_id2'
    ws['D1'] = 'Sub_id3'
    ws['E1'] = 'Sub_id4'
    ws['F1'] = 'Sub_id5'
    
    # Ghi các liên kết vào cột A, bắt đầu từ hàng số 2
    for idx, link in enumerate(links, start=2):  # Bắt đầu từ hàng 2
        ws[f'A{idx}'] = link  # Ghi vào cột A
    
    # Lưu file Excel vào bộ nhớ tạm
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

# Route để hiển thị form
@app.route('/')
def index():
    return render_template('index.html')

# Route để xử lý form và tạo file Excel
@app.route('/extract-links', methods=['POST'])
def extract_links_route():
    text = request.form['text']
    links = extract_links(text)
    
    # Nếu có file CSV, đọc các liên kết thay thế
    csv_file = request.files.get('csv_file')
    if csv_file:
        replacements = read_csv(csv_file)
        # Thay thế các liên kết gốc bằng các liên kết thay thế
        links = [replacements.get(link, link) for link in links]

    excel_file = create_excel(links)
    
    # Trả về file Excel cho người dùng
    return send_file(excel_file, as_attachment=True, download_name='links.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True)
